# -*- coding: utf-8 -*-
from collections import Counter, OrderedDict
from robot.libraries.String import String
from robot.libraries.BuiltIn import BuiltIn
from robot.api import logger
from Constant import CurrencyConstant
from Constant import CountryConstant
from SyexUiaLibrary import SyexUiaLibrary
from SyexDateTimeLibrary import SyexDateTimeLibrary
from pywinauto.application import Application
from pywinauto.controls.win32_controls import ComboBoxWrapper
from pywinauto.findbestmatch import MatchError
from uialibrary import Win32API as win32api
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_DOWN, ROUND_DOWN, ROUND_UP
import pywinauto
import os
import getpass
import math 
import re
import win32clipboard
import autoit
import time
import uialibrary
from openpyxl import load_workbook
import os.path as path
import sys
reload(sys)
sys.setdefaultencoding('Cp1252')

class SyexCustomLibrary:

    ROBOT_LIBRARY_SCOPE = 'Global'

    def __init__(self):
        self.app = Application()
        self.builtin = BuiltIn()

    def _create_app_instance(self):
        appconnect = self.app.connect(path='PowerExpress.exe')
        app = appconnect[u'WindowsForms10.Window.8.app.0.3309ded_r17_ad1']
        app.wait('ready')
        return app

    def launch_power_express(self, version, syex_env, username, use_mock_env="False"):
        self.terminate_power_express()
        express_path = self.get_power_express_path(version)
        if use_mock_env == "True":
            use_mock_env = " MockGdsLookup MockPortrait"
        else:
            use_mock_env = ""
        express_exec = 'PowerExpress.exe' + ' ENV:' + syex_env +' testuser:' + username + use_mock_env +''
        os.chdir(express_path)
        try:
            autoit.run(express_exec)
        except Exception as e:
            self.builtin.fail(str(e) + '. PowerExpress is not installed in default installation path')
        autoit.process_wait('PowerExpress.exe', timeout=240)                     
        # self._create_app_instance()
        self.builtin.set_test_variable('${username}', username)
        self.builtin.set_suite_variable('${syex_env}', syex_env)

    def terminate_power_express(self):
        win32api.TerminateProcessByName('WerFault.exe')
        win32api.TerminateProcessByName('PowerExpress.exe')
        autoit.process_wait_close('PowerExpress.exe', timeout=60)
        # os.system('TASKKILL /F /FI "USERNAME eq {}" /IM "WerFault.exe" /IM "PowerExpress.exe"'.format(self.get_local_username()))

    def get_power_express_path(self, version):
        express_basedir = "C:\\Program Files (x86)\\Carlson Wagonlit Travel\\Power Express "
        express_path = "{}{}".format(express_basedir, version)
        if os.path.exists(express_path):
            return express_path
        else:
            self.builtin.fail('Path not found. Change your Express installation to default location')

    def get_log_path(self, test_environment, version):
        local_log_path = os.getenv('LOCALAPPDATA') + "\\Carlson Wagonlit Travel\\Power Express\\"+ version +"\\"
        citrix_log_path = "D:\\Syex_Logs\\"
        if test_environment == 'citrix':
            return citrix_log_path
        else:
            return local_log_path

    def get_home_dir_path(self):
        return os.path.expanduser('~')

    def get_UserName(self):
        return self.get_local_username()

    def get_local_username(self):
        return getpass.getuser()

    def select_profile(self, user_profile):
        try:
            autoit.win_wait('[REGEXPTITLE:rofil]', 60)
            autoit.win_activate('[REGEXPTITLE:rofil]')
            select_profile_app = self.app.connect(path='PowerExpress.exe')
            select_profile_window = select_profile_app.window(title_re=u'.*rofil')
            # select_profile_window = select_profile_app.top_window_()
            select_profile_window_list_box = select_profile_window.ListBox
            select_profile_window_list_box.select(user_profile)
            select_profile_ok_button = select_profile_window.OK
            select_profile_ok_button.click()
            self.builtin.set_suite_variable('${user_profile}', user_profile)
            time.sleep(2)
        except Exception:
            pass
        time.sleep(2)
        try:
            if autoit.control_command('Power Express', '[NAME:YesBtn]', 'IsVisible'):
                autoit.control_click('Power Express', '[NAME:YesBtn]')
        except Exception:
            pass

    def select_team_name(self, team_name):
        SyexUiaLibrary().toggle_checkbox(team_name)

    def get_team_index_value(self, team_name, user_profile, syex_env='Test', window_title=None):
        team_list = [x.Name for x in uialibrary.ListControl(AutomationId='chklistTeam')\
                    .GetChildren() if x.Name != "Vertical"]
        team_name_index = team_list.index(team_name)
        logger.info("Team index is '{}'".format(team_name_index))
        return team_name_index
       
    def get_all_panels(self):
        """ 
        Returns all active / current panels

        Example:
        | @{current_panels} = | Get All Panels |

        =>
        @{current_panels} =  [ Client Info | Client Fees ]
        """
        syexuia = SyexUiaLibrary().create_power_express_handle()
        return [item.Name for item in uialibrary.TabControl(searchFromControl=syexuia).GetChildren()]

    def select_value_from_dropdown_list(self, control_id, value_to_select, window_title='Power Express', by_index=False):
        """ 
        Selects dropdown value using text or index.  If using index, use by_index=True
        
        Example:
        | Select Value From Dropdown List | [NAME:ccboClass_1] | Class Code Sample |
        """
        logger.info("Selecting dropdown value '{}'".format(value_to_select))
        if str(by_index) == "False":
            syexuia = SyexUiaLibrary()
            syexuia.select_dropdown_value(control_id, value_to_select)
        else:
            try:
                autoit.control_click(window_title, control_id)
                autoit.send("{PGUP}")
                autoit.send('{BACKSPACE}')
                autoit.send('{HOME}')
            except Exception:
                self.builtin.fail("Control '{}' not found".format(control_id))
            autoit.control_focus(window_title, control_id)
            if value_to_select == "0":
                autoit.send('{ENTER}')
                autoit.send('{TAB}')
            else:
                autoit.send("{DOWN}" * int(value_to_select))
                autoit.send('{ENTER}')
                autoit.send('{TAB}')

    def get_value_from_dropdown_list(self, control_id, window_title='Power Express'):
        """ 
        Returns a list containing all values of the specified dropdown.
        The default dropdown value is retained.
        """
        syexuia = SyexUiaLibrary() 
        return syexuia.get_dropdown_values(control_id)

    def select_value_from_combobox(self, combobox_field, combobox_value):
        """ 
        Description:
        Selects value from dropdown using dropdown name or control id as argument
        Usage:
        | Select Value From Combobox | Bundled Fee | Apply Bundled Fee |
        """
        if bool(re.search(r"(\[.*:)|(\])", combobox_field)):
            uia = SyexUiaLibrary()
            uia.select_dropdown_value(combobox_field, combobox_value)
        else:
            self._select_combobox_using_combobox_name(combobox_field, combobox_value)

    def _select_combobox_using_combobox_name(self, combobox_name, combobox_value):
        logger.info("Selecting '{}' from '{}' combobox".format(combobox_value, combobox_name))            
        try:
            self._create_app_instance()[u''.join(combobox_name)+'ComboBox'].Select(u''.join(combobox_value))
        except (ValueError):
            self.builtin.fail("Combobox value '{}' is not found".format(combobox_value))

    def get_value_from_combobox(self, combobox_name):
        """ 
        Description:
        Gets all dropdown values/items using dropdown name as argument

        Usage:
        | Get Value From Combobox | Bundled Fee |
        
        =>
        [u'Apply Bundled Fee', u'No Bundled Fee']
        """
        try:
            return self._create_app_instance()[u''.join(combobox_name)+'ComboBox'].ItemTexts()
        except Exception:
            self.builtin.fail("Usage of keyword is incorrect, try to use other keyword")

    def select_value_from_listbox(self, listbox_value):
        uia = SyexUiaLibrary().connect_to_express_via_uia()
        listbox = uia.ListBox
        listbox.Select(u''.join(listbox_value))

    def select_tst_fare(self, fare_number):
        """ 
        Description:
        For Rail Panel
        Select Tab using given EXACT value
        """
        listbox = self._create_app_instance().ListBox
        listbox.Select(u'  %s' % fare_number)

    def get_currency(self, country_code):
        return getattr(CurrencyConstant, country_code.upper())

    def get_percentage_value(self, number):
        percentage = (float(number) / float(100))
        return percentage

    def convert_to_float(self, value, precision=2):
        """ 
        Converts value to float. By default precision used is 2.

        Example:
        | ${convered_value} = | Convert To Float | 15.0909 | precision=5 |

        """
        try:
            return "{:.{prec}f}".format(float(value), prec=precision)
        except Exception as error:
            raise ValueError(repr(error))

    def get_visible_tab(self):
        """ 
        Returns fare tabs in list data type

        """
        tab_list = [tab.Name for tab in uialibrary.TabControl(AutomationId='TabControl1').GetChildren() 
                    if tab.ControlTypeName == 'TabItemControl']
        return tab_list

    def get_string_matching_regexp(self, reg_exp, data, group_number=0):
        """ 
        Get string matching Reg Expression.

        Example:
        | ${string} = | Get String Matching Regexp | TAX [0-9]+\.[0-9][0-9] | NZD120.00 TAX12.00 |
        
        Output:
        ${string} = TAX12.00
        
        """
        try:
            reg_exp = re.compile(reg_exp)
            m = re.search(reg_exp, data)
            return m.group(int(group_number))
        except Exception:
            return 0

    def get_string_using_marker(self, whole_string, first_marker, end_marker=None):
        """ 
        Description:
        Returns substring or in between string of first marker and end marker.
        Leading and trailing spaces will be stripped.

        Usage:
        | ${summary_texts} = | Get String Using Marker | HP*BIRTHDATE-20JUN84 98| - | ${SPACE} |

        =>
        ${summary_texts} = 20JUN84
        """
        self.builtin.log("Getting string from {} between {} and {}".format(whole_string, first_marker, end_marker))
        start = str(whole_string).find(first_marker) + len(first_marker)
        end = whole_string.find(end_marker, start)
        if end == -1:
            end = len(whole_string)
            return whole_string[start:end].strip()
        else:
            return whole_string[start:end].strip()      

    def get_minimum_value_from_list(self, list_):
        """ 
        Description:
        Returns the least / minimum value of list item

        Usage:
        | @{fare_list}          = | Create List                 | 12.3 | 21.2 | 35.3 | 10.8 |
        | ${minimum_fare_value} = | Get Minimum Value From List | ${fare_list}              | 

        Result:
        | ${minimum_fare_value} = | 10.8 |

        """
        try:
            return min(float(value) for value in list_)
        except ValueError:
            self.builtin.fail("List does not contain any value")

    def get_required_flight_details(self, raw_flight_details, city_pair_marker):
        raw_flight_details = raw_flight_details.strip()
        flight_details_line = raw_flight_details.find(city_pair_marker)
        fare_line_length = flight_details_line + len(city_pair_marker)
        raw_flight_details = raw_flight_details[:fare_line_length]
        flight_details_list = raw_flight_details.split(" ")
        flight_details_list = [x for x in flight_details_list if x]
        del flight_details_list[0]
        return flight_details_list

    def round_up_hk(self, number):
        """
        Description:
        Returns round up value in integer if decimal is found otherwise no roundup 
        """
        number_decimal = str(number-int(number))[2:]
        if float(number_decimal) == 0:
            return int(number)
        else:
            return int(number) + 1

    def round_off(self, number, precision=2):
        """
        Description:
        Returns round off to 2 decimal places. By default the precision is two (2)

        Usage:
        | ${number} | Round Off | 12.8998 |

        Output:
        ${number} = 12.90
        """
        try:
            number_decimal = round(float(number), int(precision))
            return str(number_decimal)
        except (ValueError):
            self.builtin.fail("'{}' is an invalid input.".format(number))

    def click_add_manual_fare(self):
        uialibrary.ButtonControl(AutomationId='btnAddManualFare').Click()

    def click_add_alternate_fare(self):
        uialibrary.ButtonControl(AutomationId='btnAddAlternateFare').Click()        

    def click_remove_fare(self):
        uialibrary.ButtonControl(AutomationId='btnRemoveFare').Click()           

    def get_tst_list(self):
        """ 
        Description:
        Get all TST item and save it to list type

        Usage:
        | @{tst_list} = | get_tst_list |

        Result:
        | @{list} =     | 1 | 2 |
        """    
        tst_listbox = self._create_app_instance().ListBox
        tst_list =  tst_listbox.ItemTexts()
        tst_list = [item.strip() for item in tst_list]
        return tst_list

    def remove_duplicate_from_list(self, list_):
        """ 
        Description:
        Remove all duplicate item from list

        Usage:
        | @{sample_list} = | Create List                | 1  | 2 | 3 | 1 |
        | @{sample_list} = | Remove Duplicate From List | ${sample_list} | 

        Result:
        | @{sample_list} = | [ 1, 2, 3 ] |
        """   
        seen = set()
        seen_add = seen.add
        return [value for value in list_ if not (value in seen or seen_add(value))]

    def replace_value_in_list(self, list_, index_or_value, new_value):
        if index_or_value.isdigit():
            try:
                list_[int(index_or_value)] = new_value
            except IndexError:
                self.builtin.fail("Index value {} is incorrect".format(index_or_value))
        else:
            try:
                index = list_.index(index_or_value)
                list_[index] = new_value
            except ValueError:
                self.builtin.fail("{} is not in list: Actual list values are : {}".\
                    format(index_or_value, ", ".join(list_)))
        return list_

    def click_menu_item(self, menu_title, double_click=False, window="Power Express"):
        """ 
        Description:
        Selects Menu item such as "Change Team", "Manual Contact", etc.
        
        Usage:
        | Click Menu Item | Change Team |

        """
        # app_dlg = self._create_power_express_backend_app()
        uia = SyexUiaLibrary().connect_to_express_via_uia()        
        menu_item = uia.child_window(title=menu_title, control_type="Button")
        if double_click == "True":
            menu_item.click_input(double=True)
        else:
            menu_item.click_input()

    def select_tab_control(self, tab_control_value, parent_tab_auto_id=None, by_index=False):
        """ 
        Description:
        Select Tab using given EXACT value
        Example:
        | Select Tab Control | Fare 1 |             |      |
        | Select Tab Control | 2      | tabCarFares | True | 

        First example selects tab by name
        Second one selects second tab (by number)
        """
        logger.info("Selecting '{}' tab".format(tab_control_value))
        syex = SyexUiaLibrary()
        syex.create_power_express_handle()
        if by_index == "True":
            tab_items = syex.get_tab_items(parent_tab_auto_id)
            tab_control_value = tab_items[int(tab_control_value) - 1]
        try:
            syex.click_tabitem_control(tab_control_value, parent_tab_auto_id)      
        except LookupError:
            self.builtin.fail("Tab '{}' is not visible".format(tab_control_value))

    def select_panel(self, panel_name):
        """ 
        Selects Panel item
        
        Example:
        | Select Panel | Cust Refs |
        """
        logger.info("Selecting panel : {}".format(panel_name))
        syex = SyexUiaLibrary()
        syex.create_power_express_handle()
        syex.click_tabitem_control(panel_name)

    def is_panel_selected(self, panel_name):
        try:
            syexuia = uialibrary.WindowControl(AutomationId='frmMain')
            return uialibrary.TabItemControl(Name=panel_name, searchFromControl=syexuia).CurrentIsSelected()
        except LookupError:
            self.builtin.fail('Panel name "{}" not found'.format(panel_name))

    def is_tab_visible(self, tab_value):
        tab_control_texts = self._create_app_instance().TabControl.Texts()
        return True if tab_value in tab_control_texts else False

    def get_lines_using_regexp(self, string, regex_pattern):
        """ 
        Description:
        Returns lines matching regex or pattern

        Usage:
        | Get Lines Using Regexp | ${pnr_details} | (QU|QA)/QE/PARWL24CA/70C4 |
        """
        regex_pattern = re.compile(regex_pattern)
        logger.info("Searching pattern: '{}'".format(regex_pattern.pattern))
        line_match = [line.strip() for line in string.splitlines() if re.search(regex_pattern, line)]
        return '\n'.join(line_match)

    def flatten_string(self, string_to_flatten, remove_space=False, remove_gds_number=False):
        """ 
        Description:
        Converts multiple lines into one line 

        Usage:
        | ${sample} | Flatten String | ${string_to_flatten} |
        
        """
        if remove_gds_number:
            string_to_flatten = "\n ".join([re.sub(r'^\s*?\d+\s', '', line) for line in string_to_flatten.splitlines()])
        
        string_to_flatten = re.sub(r"([\n ])\1*", r"\1", string_to_flatten)
        if remove_space:
            flattened_string = [line.strip() for line in string_to_flatten.splitlines()]
        else:
            flattened_string = [line.rstrip() for line in string_to_flatten.splitlines()]
        print("".join(flattened_string))
        return "".join(flattened_string)

    def verify_list_values_are_identical(self, list_):
        """ 
        Description:
        Returns Boolean value if list contains similar value

        Usage:
        | ${is_similar_value} | Verify List Values Are Identical  | ${list_sample} |
        
        """
        if list_: 
            return len(set(list_)) <= 1
        else:
            self.builtin.fail("List does not contain any value")

    def get_summary_texts(self):
        """ 
        Description:
        Returns summary text in list data type

        Usage:
        | ${summary_texts} = | Get Summary Texts  |
        
        """        
        # appwindow = self._create_power_express_backend_app()
        uia = SyexUiaLibrary().connect_to_express_via_uia()        
        summary_field = uia.child_window(auto_id="lvwSummary", control_type="List")
        summary_texts = [item for sublist in summary_field.texts() for item in sublist]
        return summary_texts

    def get_document_items(self):
        """ 
        Description:
        Retrieve all information from Unused Document tab

        Usage:
        | ${actual_document_items} = | Get Document Items |
        """    
        treeview_dlg = self._create_app_instance().TreeView
        return treeview_dlg.print_items().strip()

    def click_document_item(self, document_item, double_click=False):
        """ 
        Description:
        Clicks item from Unused Document

        Usage:
        | Click Document Item | PA S 56TGHY  079 1212121212                  USD3434     20180606  FULL | double_click=False |
        """
        treeview_dlg = self._create_app_instance().TreeView
        try:
            tree_item = treeview_dlg.get_item([u''.join(document_item)])
            logger.info("selecting '{}'".format(document_item))
            if double_click == "True":
                tree_item.click_input(double=True)
            else:
                tree_item.click_input()
        except (IndexError):
            self.builtin.fail("Document item '{}' is not found and cannot be selected".format(document_item))

    def get_service_option_items(self, apply_dict=False):
        """ 
        Description:
        Gets all values of service option delimited by Tab and every item is splitted into lines
        Returns value in list

        Usage:
        | ${service_option_values}= | Get Service Option Items |

        =>
        High Fare Calculation   WPNC‡XR¦1S  High Fare
        Low Fare International Calculation  WPNIN/T2¦1S Low Fare
        """
        list_items = self._create_app_instance().list_view.texts()
        so_list = list_items[1:]
        # print(so_list)
        if apply_dict == "True":
            service_options = []
            for index, _ in enumerate(so_list):
                 if index % 3 == 0:
                    service_options.append({so_list[index] : so_list[index + 1]})
        else:
            service_options = [so_list[index: index+2] for index, _ in enumerate(so_list)\
                     if index % 3 == 0]
        return service_options

    def get_fare_restiction_selected_radio_button(self, translate_to_english=True):
        """ 
        Description:
        Gets the Fare Rectrion name currently selected radio button
        Add argument True if translation to English is needed otherwise will return current language

        Usage:
        | ${currently_selected_fare_restriction} | Get Fare Restiction Selected Radio Button |
        | ${currently_selected_fare_restriction} | Get Fare Restiction Selected Radio Button | True |

        """        
        uia = SyexUiaLibrary().connect_to_express_via_uia()        
        restriction_value = ""
        for restriction in uia.child_window(auto_id="Panel1", control_type="Pane", found_index=0).children_texts():
            radio_button_state = uia.window(title=restriction,  control_type="RadioButton").is_selected()
            if radio_button_state == 1:
                restriction_value += restriction
                break
        return self._translate_fare_restriction(restriction_value, translate_to_english)    

    def _translate_fare_restriction(self, restriction_value, translate_to_english):
        new_restriction_value = ""
        if translate_to_english == "True":
            if restriction_value == "Modifiable":
                new_restriction_value += "Fully Flexible"
            elif restriction_value == "Partiellement modifiable":
                new_restriction_value += "Semi Flexible"
            elif restriction_value == "Non modifiable":
                new_restriction_value += "Non Flexible"
        else:
            new_restriction_value += restriction_value
        return new_restriction_value

    def truncate(self, value, precision=2):
        """
        Truncate the value without rounding up. By default precision used is 2.

        Example:
        ${truncated_value} = | Truncate | 15.0909 | precision=2

        Output:
        ${truncated_value} = 15.09
        """
        try:
            s = '{}'.format(value)
            if 'e' in s or 'E' in s:
                return '{0:.{1}value}'.format(value, precision)
            i, d = s.partition('.')
            return '.'.join([i, (d+'0'*precision)[:precision]])
        except Exception as error:
            raise ValueError(repr(error))

    def get_country_code_based_on_currency(self, currency):
        return getattr(CountryConstant, currency.lower())

    def select_gds_value(self, gds):
        logger.info("Selecting '{}' from gds toolstrip".format(gds))
        syex = SyexUiaLibrary().create_power_express_handle()
        toolstrip_button = uialibrary\
            .ButtonControl(searchFromControl=syex, RegexName='ToolStripSplitButton1|Bandeau de partage du bandeau outil1')
        toolstrip_button.SetFocus()
        toolstrip_button.Click(-10, -10)
        toolstrip_button.Click(-5, -10)
        toolstrip_button.SendKeys('{Tab}')
        gds_mapping = {'apollo':'p','amadeus':'a', 'galileo':'g', 'sabre':'s'}
        selected_gds = gds_mapping.get(gds)
        toolstrip_button.SendKeys(selected_gds)

    def get_data_from_clipboard(self):
        return uialibrary.GetClipboardText()

    def clear_data_from_clipboard(self):
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.CloseClipboard()

    def convert_list_to_lines(self, list_):
        return "\n".join(list_)

    def remove_empty_value_from_list(self, list_):
        return filter(None, list_)

    def remove_leading_and_ending_spaces(self, item):
        return item.strip()

    def remove_all_spaces(self, item):
        return item.replace(" ", "")

    def get_latest_command_response(self, gds_data, gds=None):      
        if gds == 'sabre':
            start_index = 0
            for i, v in reversed(list(enumerate(gds_data))):
                if v.startswith(">"):
                    start_index = i
                    break
            latest_response = gds_data[start_index:]
            print(latest_response)
            return self.convert_list_to_lines(latest_response)
        else:
            if gds == 'galileo':
                index_list = [i for i, line in enumerate(gds_data) if line.startswith(">") or line.startswith(')>')]
            else:
                regexp_page = re.compile(r">.*PAGE.*\d\/.*\d")
                regexp_no_more_page = re.compile(r">NO MORE PAGE AVAILABLE")
                index_list = [i for i, line in enumerate(gds_data) if line.startswith(">") and not line.startswith(">>")\
                         and not re.match(regexp_page, line) and not re.match(regexp_no_more_page, line)]
            mdr_pattern = re.compile(r"\)>mdr|>mdr|\)>|>")
            if len(index_list) > 1:
                last_two_index_list = index_list[-2:]
                latest_response = filter(lambda k: not re.match(mdr_pattern, k), gds_data[last_two_index_list[0]:])
                return self.convert_list_to_lines(latest_response)
            else:
                return self.convert_list_to_lines(filter(lambda k: not re.match(mdr_pattern, k), gds_data))

    def are_there_duplicate_remarks(self, pnr_details, gds):
        # TODO
        # implement check for other gds
        """ 
        Notes:
        Amadeus - Remarks that starts with 'RM' will be checked
        """
        remarks_list = []
        pnr_log = Counter(pnr_log.strip().lower() for pnr_log in pnr_details.split('\n') if pnr_log.strip())
        for line in pnr_log:
            if pnr_log[line] > 1:
                if gds.lower() == 'amadeus' and line != 'rir *' and line.startswith("rm"):
                    # print 'duplicate remarks found: ' + line
                    remarks_list.append(line)
        return True if len(remarks_list) > 0 else False

    def sort_pnr_details(self, pnr_details):
        """ 
        Description:
        Sorts PNR Details and removes duplicate line

        Example:
        | ${sorted_pnr_details} = | Sort Pnr Details | ${pnr_details}
        """
        remark_list = [line for line in pnr_details.splitlines()]
        # for index, remark in enumerate(remark_list):
        #     if re.match(r'\s{1,4}\d', remark):
        #         remark_list[index] = remark.lstrip()
        # seen = set()
        # seen_add = seen.add
        # sorted_pnr_details = [remark for remark in remark_list \
                            # if not (remark in seen or seen_add(remark))]
                            # or re.match(r'[ \t\t]', remark)]
        return "\n".join(remark_list)

    def verify_no_duplicate_remarks_exists(self, pnr_details, gds=None):
        """ 
        Description:
        Verifies if two consecutive lines are duplicate. Displays remark if duplicate is found
        """
        actual_remarks_list = filter(None, [re.sub(r'(^\d*\.?\s?)', '', line) for line in pnr_details.splitlines()])
        duplicate_remark_list = []
        for i, value in enumerate(actual_remarks_list):
            current_elem = value
            next_elem = actual_remarks_list[(i + 1) % len(actual_remarks_list)]
            if current_elem == next_elem:
                duplicate_remark_list.append(current_elem)

        duplicate_remark_list_length = len(duplicate_remark_list)
        if duplicate_remark_list_length > 1:
            for elem in duplicate_remark_list:
                print(elem)
        self.builtin.should_not_be_true(duplicate_remark_list_length > 1)

    def round_apac(self, number, country):
        if '.' in str(number):
            number_split = str(number).split('.')
            whole_num = number_split[0]
            decimal = number_split[1]
            if country.upper() == 'HK' or country.upper() == 'IN':
                if int(decimal[0]) > 4:
                    return int(whole_num) + 1
                else:
                    return int(whole_num)
            elif country.upper() == 'SG':
                if len(decimal) < 3:
                    return self.convert_to_float(number)
                else:
                    if decimal[2] == '5':
                        if decimal[:1] == '0':
                            return whole_num + '.' + decimal[:1] + str(int(decimal[:2]) + 1)
                        else:
                            return whole_num + '.' + str(int(decimal[:2]) + 1)
                    else:
                        return self.convert_to_float(number)
                        # return round(number, 2)
        else:
            if country.upper() == 'HK' or country.upper() == 'IN':
                return int(number)
            elif country.upper() == 'SG':
                return self.convert_to_float(number)

    def select_cell_in_data_grid_table(self, control_id, selected_index):
        """ 
        Description:
        Simulates the User Action Click on Grid Tables to select single cell. 

        (Note that the minimum selected_index value is 0)

        """
        select_row = SyexUiaLibrary().get_table_control(control_id).CustomControl(Name=selected_index)
        select_row.SetFocus()
        select_row.Click()

    def select_multiple_cells_in_data_grid_table(self, control_id, index_list):
        """ 
        Description:
        Simulates the User Action Ctrl+ Click on Grid Tables to select multiple cells.

        index_list is the list of indices to select to.  (Note that the minimum index_list value is 0)

        """
        table_control = SyexUiaLibrary().get_table_control(control_id)
        table_control.SetFocus()
        
        for i, index in enumerate(index_list):
            table_control.CustomControl(Name=str(index)).Click()
            if i == 0:
                autoit.send("{CTRLDOWN}")
        autoit.send("{CTRLUP}")

    def get_cell_value_in_data_grid_table(self, control_id, selected_index):
        """ 
        Description:
        Retrieve the value of the selected Cell in a Grid Table

        (Note that the minimum selected_index value is 0)

        """
        select_row = SyexUiaLibrary().get_table_control(control_id).CustomControl(Name=selected_index)
        return select_row.AccessibleCurrentValue()

    def click_cell_in_data_grid_pane(self, control_id, selected_row, selected_column):
        """ 
        Description:
        Simulates the User Action Click on Grid Panes to click a certain cell.

        (Note that the minimum selected_row/selected_column value is 0)

        """
        select_row = SyexUiaLibrary().get_pane_control(control_id).CustomControl(Name=selected_row)
        select_row.GetChildren()[int(selected_column)].SetFocus()
        select_row.GetChildren()[int(selected_column)].Click()

    def get_cell_value_in_data_grid_pane(self, control_id, selected_row, selected_column):
        """ 
        Description:
        Get the value of the selected Cell in a Grid Pane

        (Note that the minimum selected_row/selected_column value is 0)

        """
        select_row = SyexUiaLibrary().get_pane_control(control_id).CustomControl(Name=selected_row)
        return select_row.GetChildren()[int(selected_column)].AccessibleCurrentValue()
        
    def get_all_cell_values_in_data_grid_table(self, control_id):
        """ 
        Description:
        Retrieve the values of all cell values in a Grid Table

        """
        select_table = SyexUiaLibrary().get_table_control(control_id)
        return [x.AccessibleCurrentValue().replace('(null)', '') for x in select_table.GetChildren() if x.AccessibleCurrentValue() != "0"]

    def get_all_cell_values_in_data_grid_pane(self, control_id, column_index=2):
        """ 
        Description:
        Retrieve the values of all cell values in a Grid Pane 

        Note:
        This is specific to Grid Panes in Remarks Tab (Other Services where the texts are located at third column)

        """
        select_pane = SyexUiaLibrary().get_pane_control(control_id)
        select_pane_list = [x.GetChildren()[int(column_index)].AccessibleCurrentValue() for x in select_pane.GetChildren()]
        return filter(None, select_pane_list)

    def _get_name(self, control_id, number_of_records=5):
        # Need to be revisit the logic
        list = []
        for x in range(int(number_of_records)):
            list.append(SyexUiaLibrary().get_table_control(control_id).CustomControl(Name='Name Row ' + str(x)).AccessibleCurrentValue())    
        return list

    def delete_record_in_table_grid(self, control_field, record_to_delete, actual_number_of_records):
        # Need to be revisit the logic
        current_records = self._get_name(control_field, actual_number_of_records)
        autoit.send("{DOWN}" * int(current_records.index(record_to_delete)))
        for x in SyexUiaLibrary().get_table_control(control_field).GetChildren():
          if x.ControlTypeName == "CustomControl":
              if x.Name == "Row " + str(current_records.index(record_to_delete)):
                  for y in x.GetChildren():
                      if y.AccessibleCurrentValue() == record_to_delete:
                          y.GetPreviousSiblingControl().Click()
    
    def get_segments_from_list_control(self):
        """
        Description: Get All Segments from the Grid
        | ${segments} | Get Segments In Data Grid  |
        
        Output:
        [u'2 PR 510 Y 25FEB 1 SINMNL HK1 0020 0355']
        """     
        segments = [segment.Name for segment in uialibrary.ListControl(AutomationId='SegmentListView').GetChildren() if segment.Name != 'Header Control' and segment.Name != 'Vertical' and segment.Name != 'Horizontal' and segment.Name != '']
        return segments

    def click_amend_eo_button(self, eo_number):
        """
        Search for specified EO number and Click the Amend button
        | Click Amend EO | 1808100081  |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        try:
            row = self.get_row_object_from_datagrid(eo_number, 'EoGrid')
            row[2].SetFocus()
            row[2].Click()       
        except LookupError:
            self.builtin.fail("No EO Record.")
        # finally:
        #     ImageHorizonLibrary().take_a_screenshot()

    def delete_associated_charges_record(self, control_field, record):
        """
        Delete existing record from the grid
        | Delete Record In Associated Charges | VAT  |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        control_field = 'AssociatedCharges_AssociatedInfoGrid'
        row = self.get_row_object_from_datagrid(record, control_field)
        row[1].Click()

    def delete_vendor_grid_record(self, control_field, detail_name):
        """
        Delete existing record from the grid
        | Delete Record In Vendor Tab |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        control_field = 'ContactTypeDataGridView'
        row = self.get_row_object_from_datagrid(detail_name, control_field)
        row[1].Click()

    def delete_train_grid_record(self, control_field, train_name):
        """
        Delete existing record from the grid
        | Delete Record In Request Tab |

        """
        SyexUiaLibrary().create_power_express_handle()
        row = self.get_row_object_from_datagrid(train_name, control_field)
        row[1].Click()
        # uialibrary.ButtonControl(Name="Yes")

    def _get_eo_grid_handle(self):
        syex = SyexUiaLibrary().create_power_express_handle()
        eo_datagrid = uialibrary.TableControl(AutomationId='FromEoDataGridView', searchFromControl=syex)
        return eo_datagrid

    def _get_io_grid_handle(self):
        syex = SyexUiaLibrary().create_power_express_handle()
        io_datagrid = uialibrary.TableControl(AutomationId='FromIoDataGridView', searchFromControl=syex)
        return io_datagrid

    def get_all_remarks_from_eo_grid(self):
        eo_datagrid_children = self._get_eo_grid_handle().GetChildren()
        return [x.AccessibleCurrentValue() for x in eo_datagrid_children
                if x.ControlType == uialibrary.ControlType.CustomControl]

    def get_all_remarks_from_io_grid(self):
        io_datagrid_children = self._get_io_grid_handle().GetChildren()
        return [x.AccessibleCurrentValue() for x in io_datagrid_children
                if x.ControlType == uialibrary.ControlType.CustomControl]

    def select_remarks_from_eo_grid(self, *remarks):
        eo_remarks = self.get_all_remarks_from_eo_grid()
        eo_datagrid = self._get_eo_grid_handle()
        for remark in remarks:
            try:
                remark_index = eo_remarks.index(remark)
            except ValueError:
                self.builtin.fail("Remark '{}' not found on grid".format(remark))
            logger.info("Selecting '{}' remark from eo grid".format(remark))
            eo_datagrid.SetFocus()
            eo_datagrid.SendKeys("{CTRL}{HOME}")
            if remark_index > 0:
                eo_datagrid.SendKeys("{DOWN}" * int(remark_index))
            uialibrary.ButtonControl(AutomationId='EoAddButton').Click()
           
    def select_remarks_from_io_grid(self, *remarks):
        io_remarks = self.get_all_remarks_from_io_grid()
        io_datagrid = self._get_io_grid_handle()
        for remark in remarks:                    
            try:
                remark_index = io_remarks.index(remark)
            except ValueError:
                self.builtin.fail("Remark '{}' not found on grid".format(remark))
            logger.info("Selecting '{}' remark from io grid".format(remark))
            io_datagrid.SetFocus()
            io_datagrid.SendKeys("{CTRL}{HOME}")
            if remark_index > 0:
                io_datagrid.SendKeys("{DOWN}" * int(remark_index))
            uialibrary.ButtonControl(AutomationId='IoAddButton').Click()
            
    def click_send_email_eo(self, eo_number):
        """
        Search for specified EO number and Click the Amend button
        | Click Send Email EO | 1808100081  |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        row = self.get_row_object_from_datagrid(eo_number, 'EoGrid')
        row[2].Click()

    def convert_segment_number_to_gds_format(self):
        """
        Get All Selected Segments From Select Air Segments Grid

        Example Selected Segments
           Selected (X)     | '2 PR 510 Y 07MAR 4 SINMNL HK1 0020 0355'
           Not Selected (0) | '3 PR 300 Y 10MAR 7 MNLHKG HK1 0755 1005'
           Selected (X)     | '4 PR 313 Y 15MAR 5 HKGMNL HK1 0750 1000'
        Output:
            segment_number_short = 2,4
            segment_number_long = 0204

        Example Selected Segments
           Selected (X)     | '2 PR 510 Y 07MAR 4 SINMNL HK1 0020 0355'
           Selected (X)     | '3 PR 300 Y 10MAR 7 MNLHKG HK1 0755 1005'
           Selected (X)     | '4 PR 313 Y 15MAR 5 HKGMNL HK1 0750 1000'
        Output:
            segment_number_short = 2-4
            segment_number_long = 020304

        | ${segment_number_short} | ${segment_number_long} | Convert Segment Number To GDS Format  |
        
        """
        selected_segments = self._get_segment_selected()
        segment_numbers = [number.split(' ')[0] for number in selected_segments]
        gds_segment_number_long = ['0' + number for number in segment_numbers]
        gds_segment_number_long = ''.join(gds_segment_number_long)

        if len(selected_segments) == 1:
            gds_segment_number_short = ''.join(segment_numbers)
        else:
            is_sequential = self._is_segment_number_sequential(segment_numbers)

            if is_sequential == True:
                gds_segment_number_short = segment_numbers[0] + "-" + segment_numbers[-1]
            else:
                gds_segment_number_short = ",".join(segment_numbers)

        return gds_segment_number_short, gds_segment_number_long

    def _is_segment_number_sequential(self, _segment_numbers):
        is_sequential = False
        for i, number in enumerate(_segment_numbers, 1):
            if int(number)-1 == i:
                is_sequential = True
            else:
                is_sequential = False
                break   
        return is_sequential

    def _get_segment_selected(self):
        segment_list = SyexCustomLibrary().get_segments_from_list_control()
        segments_selected_list = []
        for segment in segment_list:
            if segment != 'Header Control':
                if SyexUiaLibrary().get_checkbox_state_from_list_control(segment):
                    segments_selected_list.append(segment)
        return segments_selected_list

    def get_equivalent_city_name(self, excel_file, city_code):
        wb = load_workbook(excel_file)
        ws = wb.worksheets[0]
        city_codes_list = [cell.value for cell in [col for col in ws['A']]]
        city_names_list = [cell.value for cell in [col for col in ws['B']]]
        city_dict = dict(zip(city_codes_list, city_names_list))
        return city_dict.get(city_code)

    def get_remarks_set_from_pnr_details(self, pnr_details, start_marker, end_marker):
        # return re.findall('('+re.escape(start_marker)+'(.*?)'+ re.escape(end_marker) + ')', 
        #     pnr_details)
        logger.info(start_marker)
        logger.info(end_marker)
        first_marker = pnr_details.index(start_marker) + len(start_marker)
        last_marker = pnr_details.index(end_marker, first_marker)
        remarks_set = start_marker +  pnr_details[first_marker:last_marker] + end_marker
        remarks_set_list = filter(None, [re.sub(r"(^\s+?(\d+\.?)\s+)", '', line) \
                                for line in remarks_set.splitlines()])
        return "\n".join(remarks_set_list)

    def _get_mco_mpd_handle(self):
        mco_mpd_grid = SyexUiaLibrary().get_pane_control('RemarksDataGridView')
        return mco_mpd_grid

    def get_all_mco_mpd_remarks_from_remarks_grid(self):
        mco_mpd_children = self._get_mco_mpd_handle().GetChildren()
        children = [child.GetLastChildControl() for child in mco_mpd_children]
        remarks = [child.AccessibleCurrentValue() for child in children]
        return remarks

    def click_list_item(self, item, double_click=True):
        """
        Description: Accepts argument as regex or literal value 
        | Click List Item | 2                          | double_click=True |
        | Click List Item | 2. PR 22FEB2019 SINMNL HK1 |                   |
        """        
        if unicode(item).isnumeric():
            segment_list = SyexCustomLibrary().get_segments_from_list_control()
            segment_numbers = [number.split(' ')[0] for number in segment_list]
            segment_index = segment_numbers.index(item)

            if str(double_click) == "True":
                uialibrary.ListItemControl(Name=segment_list[segment_index]).DoubleClick()
            else:
                uialibrary.ListItemControl(Name=segment_list[segment_index]).Click()
        else:
            if str(double_click) == "True":
                uialibrary.ListItemControl(RegexName='{}'.format(item)).DoubleClick()
            else:
                uialibrary.ListItemControl(RegexName='{}'.format(item)).Click()

    def delete_itinerary_remarks_in_air_restrictions(self, control_id, remark_value, remark_instance=None):
        """
        Description: 
        Deletes Remarks present in the Itinerary Remarks Grid In AirFare and Alternate AirFare Restriction Panel
        If remarks_instance is None, Deletes All the Remarks with Remark Name passed in Arguments.
        | Delete Itinerary Remarks In Air Restrictions | Control | Remark |                      |
        | Delete Itinerary Remarks In Air Restrictions | Control | Remark | 1                    |
        """        
        control_id = SyexUiaLibrary()._convert_control_id_to_uia_format(control_id)
        itin_remarks_table = uialibrary.PaneControl(AutomationId=control_id)
        itin_remarks_table.SetFocus()
        
        parent_match_list = []
        for item in itin_remarks_table.GetChildren():
            for row in item.GetChildren():
                if row.AccessibleCurrentValue() == remark_value:
                    parent_match_list.append(row)
        
        if remark_instance is None:
            for each_parent in parent_match_list:
                try:
                    next_sibling = each_parent.GetNextSiblingControl()
                except UnboundLocalError:
                    self.builtin.fail("Data :{} is not present in grid".format(remark_value))
                next_sibling.Click()
        else:
            parent_instance = parent_match_list[int(remark_instance) - 1]
            try:
                next_sibling = parent_instance.GetNextSiblingControl()
            except UnboundLocalError:
                self.builtin.fail("Data :{} is not present in grid".format(remark_value))
            next_sibling.Click() 
  
    def click_button_in_eo_grid(self, eo_number, country_code, eo_action):
        """
        Search for specified EO number and Click the Raise Cheque button
        | Click EO Grid Button | 1808100081  | Amend
        | Click EO Grid Button | 1808100081  | PDF
        | Click EO Grid Button | 1808100081  | Email
        | Click EO Grid Button | 1808100081  | Raise Cheque
        | Click EO Grid Button | 1808100081  | Cancel
        
        """  
        if country_code.upper() == "IN":
            button_menu_dict = {
                "amend" : 2,
            }
        else:
            button_menu_dict = {
                "amend": 1, 
                "pdf": 2, 
                "email": 3, 
                "raise cheque": 4, 
                "cancel": 5
                }

        eo_action = eo_action.lower()
        button_index = button_menu_dict.get(eo_action,6)
        self.get_row_in_eo_grid(eo_number)[button_index].Click()

    def click_button_in_eo_grid_by_product_name(self, eo_action, product_name, country_code):
        """
        Search for specified EO number and Click the Raise Cheque button
        | Click Button In Eo Grid By Product Name | pdf | 24 EMERGENCY SERVICE FEE | HK
        | Click Button In Eo Grid By Product Name | amend | MEET AND GREET SERVICE | SG

        """
        if country_code.upper() == "IN":
            button_menu_dict = {"amend": 2}
        else:
            button_menu_dict = {
                "amend": 1,
                "pdf": 2,
                "email": 3,
                "raise cheque": 4
            }

        eo_action = eo_action.lower()
        button_index = button_menu_dict.get(eo_action, 7)
        self.get_row_in_eo_grid(product_name)[button_index].Click()

    def click_cell_in_data_grid_table_by_row_and_eo_action(self, selected_row, eo_action):
        """
        Description:
        Simulates the User Action Click on Grid Panes to click a certain cell.
        Based on Row and EO_Action values supplied
        (Note that the minimum selected_row/selected_column value is 1)

        Sample:
        Click Cell In Data Grid Table By Row And Eo Action | 2 | "amend"
        Click Cell In Data Grid Table By Row And Eo Action | 3 | "pdf"
        Click Cell In Data Grid Table By Row And Eo Action | 5 | "raise cheque"
        """
        button_menu_dict = {
            "amend": 2,
            "pdf": 3,
            "email": 4,
            "raise cheque": 5
        }
        eo_action = eo_action.lower()
        button_index = button_menu_dict.get(eo_action)
        select_row = self._get_datagrid_children("EoGrid")[int(selected_row)-1]
        select_row.GetChildren()[int(button_index)].SetFocus()
        select_row.GetChildren()[int(button_index)].Click()

    def click_amend_in_eo_grid(self, row_number):
        self.click_cell_in_data_grid_table_by_row_and_eo_action(row_number, "amend")

    def click_pdf_in_eo_grid(self, row_number):
        self.click_cell_in_data_grid_table_by_row_and_eo_action(row_number, "pdf")

    def click_email_in_eo_grid(self, row_number):
        self.click_cell_in_data_grid_table_by_row_and_eo_action(row_number, "email")

    def click_raise_cheque_in_eo_grid(self, row_number):
        self.click_cell_in_data_grid_table_by_row_and_eo_action(row_number, "raise cheque")

    def get_row_data_from_eo_grid(self, selected_row):
        """
        :param selected_row: row number
        :return: contents of whole row
        """
        select_row = self._get_datagrid_children("EoGrid")[int(selected_row) - 1]
        return select_row.AccessibleCurrentValue()

    def get_all_fields_from_eo_grid(self):
        """
        :return: Returns entire grid values in list
        """
        row_values = []
        for each_child in self._get_datagrid_children("EoGrid"):
            row_values.append(each_child.AccessibleCurrentValue())
        return row_values

    def get_col_data_from_eo_grid(self, col_num):
        """
        :param col_num: Column number to be retrieved from EO Grid
        :return: List containing data from any column in EO grid

        Sample:
        Get Col Data From Eo Grid | 4 --> for EO List
        Get Col Data From Eo Grid | 5 --> for Product List
        """
        col_list = []
        grid_data = self.get_all_fields_from_eo_grid()
        eo_len = len(grid_data)
        if eo_len > 0:
            for each in grid_data:
                col = each.split(';')
                col_list.append(col[int(col_num)])
            return col_list

    def get_eo_action_status(self, row_number, eo_action):
        """
        Can be used to get status of buttons in eo grid, if it is enabled/disabled.
        :param row_number: row number for which the buttons status needs to be fetched.
        :param eo_action: "amend"/"pdf"/"email"/"raise cheque"
        :return: 'Enabled' Or 'Disabled'

        Sample:
        Get Eo Action Status | 2 | "pdf"
        Return: 'Enabled' Or 'Disabled'
        """
        button_menu_dict = {
            "amend": 1,
            "pdf": 2,
            "email": 3,
            "raise cheque": 4
        }
        eo_action = eo_action.lower()
        eo_action = button_menu_dict.get(eo_action)
        select_row = self._get_datagrid_children("EoGrid")[int(row_number) - 1]
        if select_row.GetChildren()[int(eo_action)].IsEnabled == 1:
            return "Enabled"
        else:
            return "Disabled"

    def _get_datagrid_children(self, grid_id):
        grid_parent = SyexUiaLibrary().get_table_control(grid_id).GetChildren()
        return [x for x in grid_parent if x.ControlType == uialibrary.ControlType.CustomControl][1:]

    def get_row_object_from_datagrid(self, item=None, grid_id=None):
        if item is not None:
            for selected_row in self._get_datagrid_children(grid_id):
                for column_name in selected_row.GetChildren():
                    if column_name.AccessibleCurrentValue() == item:
                        row_object = selected_row
            return row_object.GetChildren()
        else:
            return [x.AccessibleCurrentValue() for row in self._get_datagrid_children(grid_id) \
                for x in row.GetChildren() if x.ControlType == uialibrary.ControlType.CustomControl]

    def get_all_values_from_datagrid(self, grid_id):
        """
        Get all values and store it in list type
        """
        return filter(lambda l: 'System.Drawing.Bitmap' not in l, self.get_row_object_from_datagrid(grid_id=grid_id))

    def get_row_in_eo_grid(self, eo_number):
        return self.get_row_object_from_datagrid(eo_number, 'EoGrid')

    def get_row_values_from_datagrid(self, item, datagrid_id):
        datagrid_row = self.get_row_object_from_datagrid(item, datagrid_id)
        return filter(lambda x: 'System.Drawing.Bitmap' not in x, [v.AccessibleCurrentValue().\
                replace('(null)', '') for v in datagrid_row])

    def get_row_values_in_eo_grid(self, eo_number):
        return self.get_row_values_from_datagrid(eo_number, 'EoGrid')

    def get_row_fields_from_custom_fields(self):
        row_values = []
        for each_child in self._get_datagrid_children("CustomFieldGrid"):
            row_values.append(each_child.AccessibleCurrentValue())
        return row_values

    def set_row_object_in_datagrid(self, row_item, row_value, grid_auto_id, row_item_index=1, row_value_index=2):
        """
        Sets value of each datagrid in custom fields
        Sample: Set Row Object In Datagrid  <Row name from UI> <Value to be Set> <Control Id name>
        """
        for selected_row in self._get_datagrid_children(grid_auto_id):
            if selected_row.GetChildren()[int(row_item_index)].AccessibleCurrentValue() == row_item:
                selected_row.GetChildren()[int(row_value_index)].SetFocus()
                selected_row.GetChildren()[int(row_value_index)].Click()
                autoit.send('{CTRLDOWN}A{CTRLUP}')
                autoit.send(row_value)
                autoit.send('{ENTER}')

    def get_row_object_in_datagrid(self, row_item, grid_auto_id, row_item_index=1, row_value_index=2):
        """
        Sets value of each datagrid in custom fields
        Sample: Set Row Object In Datagrid  <Row name from UI> <Value to be Set> <Control Id name>
        """
        for selected_row in self._get_datagrid_children(grid_auto_id):
            if selected_row.GetChildren()[int(row_item_index)].AccessibleCurrentValue() == row_item:
                return selected_row.GetChildren()[int(row_value_index)].AccessibleCurrentValue()
                break

    def round_to_nearest_dollar(self, amount, country, round_type=None):
        """
        If round type is not specified, it will be rounded according to standard
        If country is SG, round value will return float else returns integer
        | ${round_down_value}| round to nearest dollar | 123.34 | SG | down |
        | ${round_up_value}  | round to nearest dollar | 123.34 | HK | up   |
        | ${round_value}     | round to nearest dollar | 123.34 | IN |      |

        >>
        ${round_down_value} = 2334.00
        ${round_up_value} = 2335.00
        ${round_value} = 2334.00
        """          
        round_type_dict= {'up':'ROUND_UP', 'down':'ROUND_DOWN'}
        amount_in_decimal = Decimal(amount)
        round_type = round_type_dict.get(str(round_type).lower(), 'ROUND_HALF_UP')
        round_value =  amount_in_decimal.quantize(Decimal('1'), rounding=round_type).quantize(Decimal('0.01'))
        return round_value if country.upper() == 'SG' else int(round_value)

    def get_row_index_in_error_eo_grid(self):
        error_index = []
        error_list = self.get_col_data_from_eo_grid(0)
        count = 0
        for each in error_list:
            count = count + 1
            if each =="(Icon)":
                error_index.append(count)
        return error_index

    def get_tooltip_text_in_eo_grid(self, eo_number, column_number):
        field = self.get_row_in_eo_grid(eo_number)
        field[int(column_number)].SetFocus()
        left, top, right, bottom = field[int(column_number)].BoundingRectangle
        width = right - left
        height = bottom - top
        x = left + int(float(width)/2.)
        y = top + int(float(height)/2.)
        pywinauto.mouse.move(coords=(int(x), int(y)))
        time.sleep(2)
        return uialibrary.ToolTipControl().Name

    def read_excel_files(self, excel_file, sheet_name):
        wb = load_workbook(excel_file)
        sheet_cells = []
        ws = wb.get_sheet_by_name(sheet_name)
        for row in ws.iter_rows(min_row=2):
            row_cells = []
            for cell in row:
                cell_value = cell.value
                if type(cell_value) is long:
                    cell_value = int(cell_value)
                if cell_value == 'NULL':
                    cell_value = None
                row_cells.append(cell_value)
            sheet_cells.append(tuple(row_cells))
        return sheet_cells

    def round_up_fares(self, number, round_off_to_the_nearest_value, country):
        """
        This keyword will get the next digits of the given
        round_off_to_the_nearest_value/number place value. If digits are all zero,
        ROUND_HALF_DOWN, else if there is digit other than zero, ROUND UP.
        Note: If whole number, 2 decimal digits are considered.

        Ex:
            round_off_to_the_nearest_value = 0.1
            number = 12345.876
            digit to compare = 76 = ROUND_UP = 123456.9

            round_off_to_the_nearest_value = 10
            number = 12340.678
            digit to compare = 0 and 67 = ROUND_UP = 12350

            round_off_to_the_nearest_value = 10
            number = 12340.008
            digit to compare = 0 and 00 = ROUND_HALF_DOWN = 12340

        If country is SG, round value will return float else returns integer
        | ${round_half_sg} | round up fares | 5110.8001 | 0.1 | SG | == | ${round_half_sg} = 5110.80 |
        | ${round_up_sg} | round up fares | 5110.830 | 0.1 | SG | == | ${round_up_sg} = 5110.90 |
        | ${round_half_sg2} | round up fares | 5110.803 | 1 | SG | == | ${round_half_sg2} = 5111.00 |
        | ${round_up_sg2} | round up fares | 5110.013 | 1 | SG | == | ${round_up_sg2} = 5111.00 |
        | ${round_half_hk} | round up fares | 5110.003 | 10 | HK | == | ${round_half_hk} = 5110 |
        | ${round_up_hk} | round up fares | 5113.813 | 10 | HK | == | ${round_up_hk} = 5120 |
        | ${zero_sg} | round up fares | 5113.813 | 0 | SG | == | ${zero_sg} = 5113.81 |
        | ${zero_hk} | round up fares | 0 | 10 | HK | == | ${zero_hk} = 0 |
        """
        round_off_value = float(round_off_to_the_nearest_value)
        try:
            decimal = (str(float(number)).split("."))[1]
            roundbase_decimal = str(int(float(number)/float(round_off_to_the_nearest_value)*10))[-1:]
            whole_num = (str(float(number)).split("."))[0]
            roundbase_whole = (str(float(whole_num)/float(round_off_to_the_nearest_value)).split("."))[1]
            number = math.floor(float(number) * 10 ** 2) / 10 ** 2
        except Exception:
            roundbase_decimal = 0
            roundbase_whole  = 0

        if float(round_off_to_the_nearest_value) == 0.0 or float(number) == 0.0:
            converted_value = number
        elif float(round_off_to_the_nearest_value) <= 1.00 and float(roundbase_decimal) == 0.00:
            converted_value = Decimal(str(number)).quantize(Decimal(str(round_off_value)), rounding=ROUND_DOWN)
        elif float(decimal[:2]) != 0.00 or (float(roundbase_whole) != 0.00 and float(round_off_to_the_nearest_value) > 1.00):
            converted_value = math.ceil(float(number)/float(round_off_to_the_nearest_value))*float(round_off_to_the_nearest_value)
        else:
            converted_value = float(whole_num)/float(round_off_to_the_nearest_value)*float(round_off_to_the_nearest_value)

        return self.round_apac(float(converted_value), country)
        
    def populate_other_services_custom_fields_control(self, ticket_type='ETICKET', id_type='LICENSE', id_issuing='CWT INDIA',id_number='143',no_of_passengers='2',travel_within='YES',age='16',gender='F',travel_desk_email='automation@test.com'):
        """
        This keyword will populate all visible field name in the custom fields with default values:

            'Ticket Type':'ETICKET',
            'ID Type':'LICENSE', 
            'ID Issuing Authority':'CWT INDIA',
            'ID Number':'143',
            'Number Of Passengers':'2',
            'Travel Within 24 Hours':'YES',
            'Age':'16',
            'Gender':'F',
            'Travel Desk Email ID':'automation@test.com'

        """
        dictionary = {
            'Ticket Type': ticket_type,
            'ID Type': id_type, 
            'ID Issuing Authority': id_issuing,
            'ID Number': id_number,
            'Number Of Passengers': no_of_passengers,
            'Travel Within 24 Hours': travel_within,
            'Age': age,
            'Gender': gender,
            'Travel Desk Email ID': travel_desk_email
            }
        for selected_row in self._get_datagrid_children("CustomFieldGrid"):
            selected_row.GetChildren()[2].SetFocus()
            selected_row.GetChildren()[2].Click()
            field_name = selected_row.GetChildren()[1].AccessibleCurrentValue()
            autoit.send(dictionary[field_name])
            selected_row.GetChildren()[2].SendKeys('{ENTER}')
        
    def generate_client_mi_remarks(self):
        """
        This keyword will get all values in GDS Format and MIS Value column
        And Returns Client MI Remarks Data.

        Given Table
        ... MIS Value | .... | GDS Format
        ...    A      | .... |   FF13
        ...    B      | .... |   FF14
        Output:
        string = FF13-A/FF14-B
        """
        mi_remarks = []
        for selected_row in self._get_datagrid_children("ClientMIGridView"):
            mi_remarks.append(selected_row.GetChildren()[6].AccessibleCurrentValue()+"-"+selected_row.GetChildren()[2].AccessibleCurrentValue().upper())
        return "/".join(mi_remarks)
  
    def get_tree_list(self, control_id):
        """
        This keyword retrieves the children of a Tree Object in the UI.
        This does not support nested tree objects.
        """
        tree_list =  SyexUiaLibrary().get_tree_control(control_id).GetChildren()
        tree_list = (tree_item for tree_item in tree_list if tree_item.Name != "" and tree_item.Name != "Vertical" and tree_item.Name !="Horizontal")
        tree_final_list = []
        for tree_item in tree_list:
           tree_final_list.append(tree_item.Name)
        return tree_final_list
        
    def get_eo_number_with_error_icon(self):
        error_icon_column_data = self.get_all_fields_from_eo_grid()
        eo_numbers = []
        for data in error_icon_column_data:
            data = data.split(';')
            if data[int(0)] == "(Icon)":
                eo_numbers.append(data[int(5)])
        return eo_numbers

    def replace_value_in_list_using_regex(self, pattern, replacement, list):
        """
        ${list} = ["asd n", "ttty       y"]
        Example:
        | ${list} = | Replace Value In List Using Regex | \s{2,} | ${SPACE} | ${list}

        Output
        ${list} = ["asd n", "ttty y"]
        """
        return [re.sub(pattern, replacement, x) for x in list]

    def select_car_segment(self, car_segment, by_index=False):
        """ 
        NOTE: This is for Other Services - Car Products
        
        Selects car segment in the dropdown value.
        
        Given Dropdown Values :
            2 CAR ZE HK1 JFK 16AUG 17AUG ICAR
            3 CAR 1A HK1 HKG 17OCT-18OCT CCMR

        Keyword Usage When Index is False:
        | Select Car Segment | 3 CAR 1A HK1 HKG |
        | Select Car Segment | 3 CAR 1A HK1 HKG 17OCT-18OCT CCMR |

        Keyword Usage When Index is True:
        | Select Car Segment | 1 | True
        | Select Car Segment | 2 | True
        """
        if by_index == False:
            dropdown_values = SyexUiaLibrary().get_dropdown_values('[NAME:CarSegmentsComboBox]')       
            car_segment = "".join([segment for segment in dropdown_values if car_segment in segment])

        self.select_value_from_dropdown_list('[NAME:CarSegmentsComboBox]', car_segment, by_index=by_index)

    def select_segment(self, segment, dropdown_id, by_index=False):
        """ 
        NOTE: This is for Other Services - Car/Hotel Products

        Selects car/hotel segment in the dropdown value.

        Given Dropdown Values :
            2 CAR ZE HK1 JFK 16AUG 17AUG ICAR
            3 CAR 1A HK1 HKG 17OCT-18OCT CCMR

        Usage Example:
        | Select Segment | 3 CAR 1A HK1 HKG | [NAME:CarSegmentsComboBox] |
        | Select Segment | 3 CAR 1A HK1 HKG 17OCT-18OCT CCMR | [NAME:CarSegmentsComboBox] |
        """
        # car_segment = [car_segment for segment in SyexUiaLibrary().get_dropdown_values('[NAME:CarSegmentsComboBox]') if car_segment in segment]
        
        if by_index == False:
            dropdown_values = SyexUiaLibrary().get_dropdown_values(dropdown_id)       
            segment = "".join([iti_segment for iti_segment in dropdown_values if segment in iti_segment])
            
        self.select_value_from_dropdown_list(dropdown_id, segment, by_index=by_index)

# if __name__ == '__main__':
#     syex = SyexCustomLibrary()